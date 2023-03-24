from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from processing import processing_data


def record_excel(count=5):
    wb = Workbook()
    wb.create_sheet(title='Лаба', index=0)
    sheet = wb['Лаба']
    sheet['A1'] = 'Банк'
    sheet['B1'] = 'Купить'
    sheet['C1'] = 'Продать'
    data = processing_data(count)
    for i in range(len(data)):
        print(data[i])
        cell = sheet.cell(row=i+2, column=1)
        cell.value = data[i].name
        cell = sheet.cell(row=i+2, column=2)
        cell.value = data[i].buy
        cell = sheet.cell(row=i+2, column=3)
        cell.value = data[i].sell
    titles = Reference(sheet, min_col=1, min_row=2, max_row=count+2)
    data_tmp = Reference(sheet, min_col=2, max_col=3,
                         min_row=1, max_row=count+2)
    chart = BarChart()
    chart.title = 'Гистограмма'
    chart.add_data(data_tmp, titles_from_data=True)
    chart.set_categories(titles)
    sheet.add_chart(chart, "E2")
    wb.save('laba.xls')


record_excel(5)
